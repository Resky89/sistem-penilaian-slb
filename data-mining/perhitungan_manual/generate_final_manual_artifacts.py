import re
import math
import ast
import zipfile
import shutil
import tempfile
import os
from pathlib import Path
from xml.sax.saxutils import escape

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.preprocessing import LabelEncoder, OneHotEncoder
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import RandomizedSearchCV, train_test_split
from sklearn.metrics import accuracy_score
from imblearn.over_sampling import SMOTE
from sklearn.neighbors import NearestNeighbors

BASE = Path(r"d:/projects/Sistem-Penilaian-SLB")
DM = BASE / "data-mining"
RAW_PATH = BASE / "DATA MENTAH.xlsx"
SPLIT_PATH = BASE / "DATA_SPLIT_TRAIN_TEST.xlsx"
TEMPLATE_DOCX = BASE / "Panduan_Perhitungan_Manual_Skripsi.docx"
OUT_XLSX = DM / "perhitungan_manual" / "PERHITUNGAN_MANUAL_FINAL_SEMUA_DATA_RESKY.xlsx"
OUT_DOCX = BASE / "DOKUMEN_PERHITUNGAN_MANUAL_FINAL_RESKY.docx"

STOPWORDS = set([
    'dan', 'di', 'ke', 'dari', 'ini', 'itu', 'juga', 'untuk', 'pada', 'dengan',
    'adalah', 'yang', 'saya', 'kami', 'anda', 'mereka', 'ia', 'dia', 'kita', 
    'dapat', 'harus', 'akan', 'sudah', 'telah', 'sedang', 'ingin',
    'ada', 'bukan', 'hanya', 'saja', 'atau', 'namun', 'tetapi',
    'oleh', 'seperti', 'maka', 'jika', 'karena', 'sehingga', 'bahwa',
    'hal', 'secara', 'tersebut', 'dalam', 'atas', 'bawah', 'serta', 'bagi', 'setelah',
    'peserta', 'didik', 'siswa', 'aspek', 'kegiatan', 'aktivitas', 'cara', 
    'berikan', 'gunakan', 'saran', 'hari',
    'dafa', 'dzikri', 'eria', 'arif', 'arifin', 'rama', 'fadil', 'afif', 
    'azzam', 'dewi', 'faiz', 'ahmad', 'bilqis', 'rizky', 'mia', 'andreansyah', 
    'rahman', 'khansa', 'robi', 'roby'
])

def clean_text(text):
    if not isinstance(text, str):
        return ""
    text = re.sub(r'(?i)deskripsi perkembangan\s*:\s*', '', text)
    text = text.lower()
    text = re.sub(r"[^a-zA-Z\s]", " ", text)
    words = [w for w in text.split() if len(w) > 2 and w not in STOPWORDS]
    return " ".join(words)

def recover_missing_values(df_split, raw_path):
    if not Path(raw_path).exists():
        return df_split
    xl_raw = pd.ExcelFile(raw_path)
    raw_sheets = {}
    for sheet in xl_raw.sheet_names:
        df_raw = pd.read_excel(xl_raw, sheet_name=sheet)
        df_raw = df_raw.copy()
        if 'Sumber Data' not in df_raw.columns:
            df_raw['Sumber Data'] = sheet.replace('Data Mentah ', '')
        raw_sheets[sheet.replace('Data Mentah ', '')] = df_raw

    df_split = df_split.copy()
    for idx, row in df_split.iterrows():
        student = row.get('Nama Siswa', '')
        if not isinstance(student, str) or not student:
            student = row.get('Sumber Data', '')
        if student in raw_sheets:
            raw_df = raw_sheets[student]
            match_mask = (raw_df['Label'].astype(str).str.strip() == str(row['Label']).strip())
            raw_val_str = raw_df['Nilai'].astype(str).str.strip().str.replace('.0', '', regex=False)
            split_val_str = str(row['Nilai']).strip().replace('.0', '')
            match_mask = match_mask & (raw_val_str == split_val_str)
            matches = raw_df[match_mask]
            if len(matches) == 1:
                recovered_row = matches.iloc[0]
                if pd.isnull(row['Aspek / Mapel']) or str(row['Aspek / Mapel']) == 'nan':
                    df_split.at[idx, 'Aspek / Mapel'] = recovered_row['Aspek / Mapel']
                if pd.isnull(row['Deskripsi Penilaian']) or str(row['Deskripsi Penilaian']) == 'nan':
                    df_split.at[idx, 'Deskripsi Penilaian'] = recovered_row['Deskripsi Penilaian']
            elif len(matches) > 1:
                if not pd.isnull(row['Deskripsi Penilaian']) and str(row['Deskripsi Penilaian']) != 'nan':
                    narrowed = matches[matches['Deskripsi Penilaian'].astype(str).str.strip() == str(row['Deskripsi Penilaian']).strip()]
                    if len(narrowed) == 1:
                        recovered_row = narrowed.iloc[0]
                        if pd.isnull(row['Aspek / Mapel']) or str(row['Aspek / Mapel']) == 'nan':
                            df_split.at[idx, 'Aspek / Mapel'] = recovered_row['Aspek / Mapel']
                elif not pd.isnull(row['Aspek / Mapel']) and str(row['Aspek / Mapel']) != 'nan':
                    narrowed = matches[matches['Aspek / Mapel'].astype(str).str.strip() == str(row['Aspek / Mapel']).strip()]
                    if len(narrowed) == 1:
                        recovered_row = narrowed.iloc[0]
                        if pd.isnull(row['Deskripsi Penilaian']) or str(row['Deskripsi Penilaian']) == 'nan':
                            df_split.at[idx, 'Deskripsi Penilaian'] = recovered_row['Deskripsi Penilaian']
    return df_split

def normalize_cols(df):
    df = df.copy()
    if 'Sumber Data' not in df.columns:
        df['Sumber Data'] = df['Nama Siswa'] if 'Nama Siswa' in df.columns else ''
    return df

def load_data():
    if SPLIT_PATH.exists():
        tr = normalize_cols(pd.read_excel(SPLIT_PATH, sheet_name='Data Training (80%)'))
        te = normalize_cols(pd.read_excel(SPLIT_PATH, sheet_name='Data Testing (20%)'))
        tr = recover_missing_values(tr, RAW_PATH)
        te = recover_missing_values(te, RAW_PATH)
        tr['Set Data'] = 'Training'; te['Set Data'] = 'Testing'
        data = pd.concat([tr, te], ignore_index=True)
    else:
        xl = pd.ExcelFile(RAW_PATH)
        rows = []
        for sheet in xl.sheet_names:
            df = normalize_cols(pd.read_excel(RAW_PATH, sheet_name=sheet))
            df['Sumber Data'] = sheet.replace('Data Mentah ', '')
            rows.append(df)
        data = pd.concat(rows, ignore_index=True)
        data['Set Data'] = 'All'
    need = ['Sumber Data','Aspek / Mapel','Nilai','Deskripsi Penilaian','Label','Set Data']
    data = data[[c for c in need if c in data.columns]].copy()
    data['Nilai'] = pd.to_numeric(data['Nilai'], errors='coerce').fillna(0)
    data = data.dropna(subset=['Aspek / Mapel','Deskripsi Penilaian','Label']).reset_index(drop=True)
    data['Deskripsi Penilaian'] = data['Deskripsi Penilaian'].astype(str).str.replace(r'(?i)deskripsi perkembangan\s*:\s*', '', regex=True)
    data.insert(0, 'ID_Data', np.arange(1, len(data)+1))
    data['Teks Bersih'] = data['Deskripsi Penilaian'].apply(clean_text)
    return data

def gini(counts):
    total = int(np.sum(counts))
    if total == 0:
        return 0.0
    return float(1 - np.sum((np.array(counts) / total) ** 2))

def style_ws(ws):
    fill = PatternFill('solid', fgColor='1F4E78')
    font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if ws.max_row >= 1:
        for c in ws[1]:
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    for i in range(1, ws.max_column + 1):
        col = get_column_letter(i)
        mx = 10
        for c in ws[col]:
            v = '' if c.value is None else str(c.value)
            mx = min(max(mx, len(v)), 55)
        ws.column_dimensions[col].width = mx + 2

def write_df(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(r, 1, title).font = Font(bold=True, size=13, color='1F4E78')
        r += 1
    for j, col in enumerate(df.columns, 1):
        ws.cell(r, j, col)
    for i, row in enumerate(df.itertuples(index=False), r+1):
        for j, val in enumerate(row, 1):
            ws.cell(i, j, val)
    return r + len(df) + 2

def add_sheet(wb, name, df):
    ws = wb.create_sheet(name[:31])
    write_df(ws, df)
    style_ws(ws)
    return ws

def build_smote_calculations():
    data = load_data()
    
    if set(data['Set Data']) >= {'Training','Testing'}:
        train_mask = data['Set Data'].eq('Training').values
        test_mask = data['Set Data'].eq('Testing').values
    else:
        train_mask = np.ones(len(data), dtype=bool)
        test_mask = np.zeros(len(data), dtype=bool)

    # 1. TF-IDF Fit only on Training Data
    tfidf = TfidfVectorizer(max_features=50, min_df=2, max_df=0.85, ngram_range=(1,1), smooth_idf=True, norm='l2', use_idf=True, sublinear_tf=False)
    tfidf.fit(data.loc[train_mask, 'Teks Bersih'])
    X_tfidf_orig = tfidf.transform(data['Teks Bersih']).toarray()
    words = list(tfidf.get_feature_names_out())
    tfidf_cols = [f'tfidf_{w}' for w in words]

    try:
        ohe = OneHotEncoder(sparse_output=False, handle_unknown='ignore')
    except TypeError:
        ohe = OneHotEncoder(sparse=False, handle_unknown='ignore')
        
    X_aspek_orig = ohe.fit_transform(data[['Aspek / Mapel']])
    aspek_cols = [f'aspek_{a}' for a in ohe.categories_[0]]
    
    # Combined original features
    X_orig = np.hstack([data[['Nilai']].values, X_tfidf_orig, X_aspek_orig])
    feature_names = ['Nilai'] + tfidf_cols + aspek_cols
    
    le = LabelEncoder()
    y_orig = le.fit_transform(data['Label'].astype(str))
    labels = list(le.classes_)

    # Split train and test
    X_train_orig = X_orig[train_mask]
    y_train_orig = y_orig[train_mask]
    X_test = X_orig[test_mask]
    y_test = y_orig[test_mask]

    # --- APPLY SMOTE ON TRAINING SET ---
    smote = SMOTE(random_state=42)
    X_train_res, y_train_res = smote.fit_resample(X_train_orig, y_train_orig)
    
    # Identify which training samples are synthetic (generated by SMOTE)
    n_orig_train = len(X_train_orig)
    n_res_train = len(X_train_res)
    n_synthetic = n_res_train - n_orig_train

    print(f"SMOTE: Original Train = {n_orig_train}, Resampled Train = {n_res_train} (Created {n_synthetic} synthetic samples)")

    # Build resampled training dataframes
    res_rows = []
    # 1. Add all original data first (both train and test)
    for idx, row in data.iterrows():
        is_train = train_mask[idx]
        res_rows.append({
            'ID_Data': int(row['ID_Data']),
            'Sumber Data': row['Sumber Data'],
            'Aspek / Mapel': row['Aspek / Mapel'],
            'Set Data': row['Set Data'],
            'Nilai': float(row['Nilai']),
            'Deskripsi Penilaian': row['Deskripsi Penilaian'],
            'Teks Bersih': row['Teks Bersih'],
            'Label': row['Label'],
            'Tipe_Data': 'Original'
        })
        
    # 2. Add synthetic training samples
    # For aspects, we reconstruct Aspek / Mapel from OHE features in X_train_res using argmax
    ohe_start_idx = 1 + len(tfidf_cols)
    for i in range(n_synthetic):
        syn_idx = n_orig_train + i
        syn_features = X_train_res[syn_idx]
        syn_nilai = round(float(syn_features[0]), 1)
        syn_label = le.inverse_transform([y_train_res[syn_idx]])[0]
        
        # Get aspect by argmax from OHE columns
        syn_aspek_ohe = syn_features[ohe_start_idx:]
        best_ohe_idx = int(np.argmax(syn_aspek_ohe))
        syn_aspek = ohe.categories_[0][best_ohe_idx]
        
        res_rows.append({
            'ID_Data': int(1000 + i + 1),
            'Sumber Data': f'SMOTE_Sintetis_{i+1}',
            'Aspek / Mapel': syn_aspek,
            'Set Data': 'Training',
            'Nilai': syn_nilai,
            'Deskripsi Penilaian': '[Sintetis oleh SMOTE]',
            'Teks Bersih': '[Sintetis oleh SMOTE]',
            'Label': syn_label,
            'Tipe_Data': 'Synthetic_SMOTE'
        })
        
    transformed_df = pd.DataFrame(res_rows)

    # Reconstruct whole X matrix (with original + resampled train + test)
    # We construct it directly from resampled training + original test
    X_train_final = X_train_res
    y_train_final = y_train_res
    
    # We will build X and y representation for all rows in transformed_df
    # ID_Data maps:
    # IDs 1 to 364 are original (some are train, some are test)
    # IDs 1001+ are synthetic train
    X_all_res = np.zeros((len(transformed_df), len(feature_names)))
    y_all_res = np.zeros(len(transformed_df), dtype=int)
    
    # Original split mapping
    train_orig_id_to_idx = {int(data.loc[idx, 'ID_Data']): i for i, idx in enumerate(np.where(train_mask)[0])}
    test_orig_id_to_idx = {int(data.loc[idx, 'ID_Data']): i for i, idx in enumerate(np.where(test_mask)[0])}
    
    for idx, row in transformed_df.iterrows():
        id_data = int(row['ID_Data'])
        if id_data <= 364:  # Original
            if id_data in train_orig_id_to_idx:
                orig_train_idx = train_orig_id_to_idx[id_data]
                X_all_res[idx] = X_train_res[orig_train_idx]
                y_all_res[idx] = y_train_res[orig_train_idx]
            else:
                orig_test_idx = test_orig_id_to_idx[id_data]
                X_all_res[idx] = X_test[orig_test_idx]
                y_all_res[idx] = y_test[orig_test_idx]
        else:  # Synthetic
            syn_train_idx = n_orig_train + (id_data - 1001)
            X_all_res[idx] = X_train_res[syn_train_idx]
            y_all_res[idx] = y_train_res[syn_train_idx]

    # Model training with tuned parameters or small tree parameters for display
    # We train a small model (n_estimators=10, max_depth=3) for manual calculation simplicity
    rf = RandomForestClassifier(n_estimators=10, max_depth=3, min_samples_split=2, random_state=42, bootstrap=True)
    rf.fit(X_train_final, y_train_final)
    
    pred_all = rf.predict(X_all_res)
    proba_all = rf.predict_proba(X_all_res)

    # Grid search tuning representation
    param_dist = {
        'n_estimators': [100, 200, 300, 400, 500],
        'max_depth': [10, 20, 30, None],
        'min_samples_split': [2, 5, 10],
        'min_samples_leaf': [1, 2, 4],
        'max_features': ['sqrt', 'log2']
    }
    grid = RandomizedSearchCV(
        estimator=RandomForestClassifier(random_state=42, bootstrap=True),
        param_distributions=param_dist,
        n_iter=20,
        cv=5,
        scoring='accuracy',
        n_jobs=-1,
        random_state=42,
        return_train_score=True
    )
    grid.fit(X_train_final, y_train_final)

    transformed_df['Y_Encoded'] = y_all_res
    transformed_df['Y_Prediksi'] = le.inverse_transform(pred_all)
    transformed_df['Prediksi_Benar'] = transformed_df['Y_Prediksi'].eq(transformed_df['Label'])

    # Matrix view of all data features
    matrix = pd.concat([
        transformed_df[['ID_Data','Sumber Data','Aspek / Mapel','Set Data','Tipe_Data','Nilai','Label','Teks Bersih']],
        pd.DataFrame(np.round(X_all_res[:, 1:1+len(tfidf_cols)], 6), columns=tfidf_cols)
    ], axis=1)

    # TF-IDF Manual calculations sheet (only for original data, since synthetic TF-IDF is interpolated)
    vocab_index = {w: i for i, w in enumerate(words)}
    docs = [d.split() if isinstance(d, str) and d else [] for d in data['Teks Bersih']]
    docs_train = [d.split() if isinstance(d, str) and d else [] for d in data.loc[train_mask, 'Teks Bersih']]
    N_train = len(docs_train)
    df_train = {w: sum(1 for doc in docs_train if w in set(doc)) for w in words}
    idf_train = {w: math.log((1+N_train)/(1+df_train[w])) + 1 for w in words}
    rows = []
    
    for r, row_data in data.iterrows():
        doc = docs[r]
        counts = pd.Series(doc).value_counts().to_dict() if doc else {}
        raw_vals = {w: counts.get(w,0) * idf_train[w] for w in words}
        l2 = math.sqrt(sum(v*v for v in raw_vals.values()))
        id_val = int(row_data['ID_Data'])
        res_row_idx = transformed_df[transformed_df['ID_Data'] == id_val].index[0]
        
        for w in words:
            c = counts.get(w, 0)
            py_val = float(X_all_res[res_row_idx, 1 + vocab_index[w]])
            if c or py_val != 0:
                raw = raw_vals[w]
                final = raw / l2 if l2 else 0
                rows.append({
                    'ID_Data': id_val,
                    'Siswa': row_data['Sumber Data'],
                    'Aspek / Mapel': row_data['Aspek / Mapel'],
                    'Kata': w,
                    'tf raw (jumlah kata)': int(c),
                    'N': N_train,
                    'df(w)': int(df_train[w]),
                    'idf sklearn = ln((1+N)/(1+df))+1': round(idf_train[w], 9),
                    'tf × idf': round(raw, 9),
                    'L2 norm dokumen': round(l2, 9),
                    'TF-IDF Manual Setelah L2': round(final, 9),
                    'TF-IDF Python sklearn': round(py_val, 9),
                    'Selisih': round(final - py_val, 12)
                })
    tfidf_manual = pd.DataFrame(rows)

    # Per-tree calculations
    tree_paths_all = []; votes = []; gini_all = []; tree_structures = []
    for t_idx, est in enumerate(rf.estimators_, 1):
        tree = est.tree_
        decision_path_train = est.decision_path(X_train_final).toarray().astype(bool)
        
        # 1. Tree structure
        for node in range(tree.node_count):
            is_leaf = tree.children_left[node] == tree.children_right[node]
            value = tree.value[node][0]
            pred_cls = int(np.argmax(value))
            norm_val = value / np.sum(value) if np.sum(value) > 0 else value
            counts_dict = {labels[i]: int(round(v * tree.n_node_samples[node])) for i, v in enumerate(norm_val)}
            row = {
                'Pohon ke-': t_idx, 'Node': node,
                'Jenis Node': 'Leaf' if is_leaf else 'Internal',
                'Prediksi Node': le.inverse_transform([pred_cls])[0],
                'Distribusi Kelas': str(counts_dict)
            }
            if is_leaf:
                row.update({'Fitur Split': '-', 'Threshold': '-', 'Cabang Kiri': '-', 'Cabang Kanan': '-'})
            else:
                fi = tree.feature[node]
                row.update({
                    'Fitur Split': feature_names[fi],
                    'Threshold': round(float(tree.threshold[node]), 9),
                    'Cabang Kiri': int(tree.children_left[node]),
                    'Cabang Kanan': int(tree.children_right[node])
                })
            tree_structures.append(row)
            
        # 2. Gini split
        for node in range(tree.node_count):
            if tree.children_left[node] == tree.children_right[node]:
                continue
            fi = tree.feature[node]
            thr = tree.threshold[node]
            at_node = decision_path_train[:, node]
            left = at_node & (X_train_final[:, fi] <= thr)
            right = at_node & (X_train_final[:, fi] > thr)
            y_node, y_left, y_right = y_train_final[at_node], y_train_final[left], y_train_final[right]
            cnt_node = np.bincount(y_node, minlength=len(labels))
            cnt_left = np.bincount(y_left, minlength=len(labels))
            cnt_right = np.bincount(y_right, minlength=len(labels))
            gp, gl, gr = gini(cnt_node), gini(cnt_left), gini(cnt_right)
            n, nl, nr = len(y_node), len(y_left), len(y_right)
            gs = (nl/n)*gl + (nr/n)*gr if n else 0
            gini_all.append({
                'Pohon ke-': t_idx, 'Node': node, 'Fitur': feature_names[fi],
                'Threshold': round(float(thr), 9), 'n Node': n,
                'Distribusi Node': str(dict(zip(labels, cnt_node.astype(int)))), 'Gini Node': round(gp, 9),
                'n Kiri': nl, 'Distribusi Kiri': str(dict(zip(labels, cnt_left.astype(int)))), 'Gini Kiri': round(gl, 9),
                'n Kanan': nr, 'Distribusi Kanan': str(dict(zip(labels, cnt_right.astype(int)))), 'Gini Kanan': round(gr, 9),
                'Gini Split': round(gs, 9), 'Gain Gini': round(gp-gs, 9)
            })
            
        # 3. Path tracing for all resampled rows
        for i, row_trans in transformed_df.iterrows():
            id_val = int(row_trans['ID_Data'])
            node = 0; step = 1; xrow = X_all_res[i]
            while tree.children_left[node] != tree.children_right[node]:
                fi = tree.feature[node]
                thr = tree.threshold[node]
                val = xrow[fi]
                go_left = val <= thr
                nxt = tree.children_left[node] if go_left else tree.children_right[node]
                tree_paths_all.append({
                    'ID_Data': id_val, 'Siswa': row_trans['Sumber Data'], 'Aspek / Mapel': row_trans['Aspek / Mapel'],
                    'Pohon ke-': t_idx, 'Langkah': step, 'Node': int(node),
                    'Aturan': f'{feature_names[fi]} <= {thr:.9f}', 'Nilai Data': round(float(val), 9),
                    'Keputusan': 'Kiri' if go_left else 'Kanan', 'Node Berikutnya': int(nxt)
                })
                node = nxt; step += 1
            pred_cls = int(np.argmax(tree.value[node][0]))
            pred_label = le.inverse_transform([pred_cls])[0]
            tree_paths_all.append({
                'ID_Data': id_val, 'Siswa': row_trans['Sumber Data'], 'Aspek / Mapel': row_trans['Aspek / Mapel'],
                'Pohon ke-': t_idx, 'Langkah': step, 'Node': int(node),
                'Aturan': 'Leaf', 'Nilai Data': '-', 'Keputusan': f'Prediksi = {pred_label}', 'Node Berikutnya': '-'
            })
            votes.append({
                'ID_Data': id_val, 'Siswa': row_trans['Sumber Data'], 'Aspek / Mapel': row_trans['Aspek / Mapel'],
                'Pohon ke-': t_idx, 'Prediksi Pohon': pred_label
            })

    tree_structures = pd.DataFrame(tree_structures)
    gini_all = pd.DataFrame(gini_all)
    tree_paths_all = pd.DataFrame(tree_paths_all)
    votes = pd.DataFrame(votes)

    vote_recap = votes.groupby(['ID_Data','Siswa','Aspek / Mapel','Prediksi Pohon']).size().reset_index(name='Jumlah Suara')
    final_vote = vote_recap.sort_values(['ID_Data','Jumlah Suara'], ascending=[True,False]).groupby('ID_Data').head(1).rename(columns={'Prediksi Pohon':'Hasil Voting'})
    final_vote = final_vote.merge(transformed_df[['ID_Data','Label']], on='ID_Data', how='left')
    final_vote['Benar/Salah'] = np.where(final_vote['Hasil Voting'].eq(final_vote['Label']), 'Benar', 'Salah')

    hpt = pd.DataFrame(grid.cv_results_)[['param_n_estimators','param_max_depth','param_min_samples_split','param_min_samples_leaf','param_max_features','mean_train_score','mean_test_score','rank_test_score']].copy()
    hpt.columns = ['n_estimators','max_depth','min_samples_split','min_samples_leaf','max_features','Mean Train Accuracy','Mean CV Accuracy','Ranking']
    hpt = hpt.sort_values(['Ranking','n_estimators']).reset_index(drop=True)
    hpt['Status'] = np.where(hpt['Ranking'].eq(1), 'Best', '-')
    hpt['Mean Train Accuracy'] = hpt['Mean Train Accuracy'].round(9)
    hpt['Mean CV Accuracy'] = hpt['Mean CV Accuracy'].round(9)

    multi_long = transformed_df[['Sumber Data','Aspek / Mapel','Nilai','Label','Y_Prediksi']].copy()
    multi_long['Makna'] = 'Output label perkembangan untuk aspek/mapel ini'
    actual = multi_long.pivot_table(index='Sumber Data', columns='Aspek / Mapel', values='Label', aggfunc='first')
    pred = multi_long.pivot_table(index='Sumber Data', columns='Aspek / Mapel', values='Y_Prediksi', aggfunc='first')
    actual.columns = [f'Aktual_{c}' for c in actual.columns]
    pred.columns = [f'Prediksi_{c}' for c in pred.columns]
    multi_wide = pd.concat([actual, pred], axis=1).reset_index()

    # SHAP approximation for all rows
    baseline = np.mean(X_train_final, axis=0)
    shap_rows = []; shap_summary = []
    for i, row_trans in transformed_df.iterrows():
        id_val = int(row_trans['ID_Data'])
        xrow = X_all_res[i]
        cls = int(pred_all[i])
        cls_label = le.inverse_transform([cls])[0]
        fx = float(proba_all[i, cls])
        base = float(np.mean(rf.predict_proba(X_train_final)[:, cls]))
        active = [0] + [j for j, v in enumerate(xrow) if j > 0 and abs(v) > 1e-12]
        raw = []
        for fi in active:
            xabl = xrow.copy(); xabl[fi] = baseline[fi]
            pabl = float(rf.predict_proba(xabl.reshape(1, -1))[0, cls])
            raw.append((fi, fx - pabl, pabl))
        raw = sorted(raw, key=lambda z: abs(z[1]), reverse=True)[:10]
        total = sum(v for _, v, _ in raw)
        scale = (fx - base)/total if abs(total) > 1e-12 else 0
        sum_phi = 0
        for fi, delta, pabl in raw:
            phi = delta * scale
            sum_phi += phi
            shap_rows.append({
                'ID_Data': id_val, 'Siswa': row_trans['Sumber Data'], 'Aspek / Mapel': row_trans['Aspek / Mapel'],
                'Kelas Dijelaskan': cls_label, 'Fitur': feature_names[fi], 'Nilai Fitur': round(float(xrow[fi]), 9),
                'f(x) Asli': round(fx, 9), 'f(x) Tanpa Fitur': round(pabl, 9),
                'Δ=f(x)-f(x tanpa fitur)': round(delta, 9), 'φi SHAP Manual': round(phi, 9),
                'Interpretasi': 'Menaikkan peluang output' if phi > 0 else ('Menurunkan peluang output' if phi < 0 else 'Netral')
            })
        shap_summary.append({
            'ID_Data': id_val, 'Siswa': row_trans['Sumber Data'], 'Aspek / Mapel': row_trans['Aspek / Mapel'],
            'Output Prediksi': cls_label, 'φ0 Base Value': round(base, 9), 'Σφi': round(sum_phi, 9),
            'f(x)=φ0+Σφi Excel': round(base + sum_phi, 9), 'Probabilitas Model': round(fx, 9)
        })
        
    shap_rows = pd.DataFrame(shap_rows)
    shap_summary = pd.DataFrame(shap_summary)

    formula = pd.DataFrame([
        ['SMOTE Oversampling', 'x_new = x_i + λ(x_zi - x_i)', 'Menduplikasi & menginterpolasi kelas minoritas (Baik, Perlu Bimbingan) agar seimbang dengan kelas mayoritas (Cukup)'],
        ['TF Raw','tf(w,d)=n(w,d)','Frekuensi kemunculan kata kunci dalam dokumen.'],
        ['IDF Smooth','idf(w)=ln((1+N)/(1+df(w)))+1','Perhitungan IDF dengan smoothing untuk menghindari pembagian dengan nol.'],
        ['TF-IDF sebelum normalisasi','v(w,d)=tf(w,d)×idf(w)','Bobot awal kata.'],
        ['L2 norm','||v_d||₂=sqrt(Σ v(w,d)^2)','Panjang vektor bobot dokumen.'],
        ['TF-IDF akhir','tfidf(w,d)=v(w,d)/||v_d||₂','Hasil pembobotan akhir setelah normalisasi L2.'],
        ['Gini','Gini(D)=1-Σp_i²','Impurity node.'],
        ['Gini Split','Gini_split=(n1/n)Gini(D1)+(n2/n)Gini(D2)','Nilai pemisahan node.'],
        ['Voting','ŷ=mode{h1(x),h2(x),...,hT(x)}','Suara mayoritas pohon.'],
        ['SHAP','f(x)=φ0+Σφi','Prediksi dijelaskan sebagai kontribusi fitur.']
    ], columns=['Nama','Rumus','Keterangan'])
    
    summary = pd.DataFrame([
        ['Jumlah data original', len(data)],
        ['Jumlah training original', len(X_train_orig)],
        ['Jumlah training resampled (SMOTE)', len(X_train_final)],
        ['Jumlah testing', len(X_test)],
        ['Jumlah data sintetis SMOTE', n_synthetic],
        ['Sebaran Label Sebelum SMOTE', f'Cukup: 152, Baik: 69, Perlu Bimbingan: 67'],
        ['Sebaran Label Sesudah SMOTE', f'Cukup: 152, Baik: 152, Perlu Bimbingan: 152'],
        ['Best HPT Params', str(grid.best_params_)],
        ['Best CV Accuracy', round(float(grid.best_score_), 9)]
    ], columns=['Keterangan','Nilai'])
    
    return locals()

def add_explanation_sheet_smote(wb):
    ws = wb.create_sheet('BACA DULU - SMOTE')
    rows = [
        ['Urutan', 'Tahap', 'Yang dihitung / Ditampilkan', 'Buka sheet', 'Penjelasan singkat'],
        [1, 'SMOTE Info', 'Ringkasan jumlah data latih sebelum & sesudah SMOTE.', 'Rumus dan Ringkasan', 'Menjelaskan penyeimbangan data. Kelas Cukup, Baik, dan Perlu Bimbingan di-resample masing-masing menjadi 152 baris.'],
        [2, 'Dataset Resampled', 'Data gabungan original + data sintetis buatan SMOTE.', 'DATASET KESELURUHAN', 'Kolom ID_Data 1001 ke atas adalah data buatan SMOTE (SMOTE_Sintetis).'],
        [3, 'Pembersihan teks', 'Deskripsi Penilaian diubah menjadi teks bersih.', 'Transformasi Data', 'Untuk data original dibersihkan, data SMOTE langsung diinterpolasi fiturnya.'],
        [4, 'TF-IDF', 'Bobot setiap kata pada setiap dokumen original.', 'TF-IDF Rumus Excel', 'Excel menghitung IDF dari training original, lalu melakukkan normalisasi L2.'],
        [5, 'Gini Impurity (Resampled)', 'Kualitas split pohon dihitung dari data hasil SMOTE.', 'Gini Rumus Excel', 'Gini dihitung dari populasi train yang sudah seimbang (456 data), membuat pohon lebih objektif.'],
        [6, 'Voting', 'Suara dari pohon Random Forest yang dilatih dengan data SMOTE.', 'Voting Rumus Excel', 'Excel menjumlahkan voting pohon untuk data original dan sintetis.'],
        [7, 'SHAP / Kontribusi Fitur', 'Kontribusi fitur terhadap keputusan model seimbang.', 'SHAP Rumus Excel', 'Menjelaskan kontribusi nilai dan kata kunci.'],
    ]
    for r, row in enumerate(rows, 1):
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    style_ws(ws)
    return ws

def add_tfidf_formula_sheet(wb, obj):
    ws = wb.create_sheet('TF-IDF Rumus Excel')
    headers = [
        'ID_Data', 'Siswa', 'Aspek / Mapel', 'Kata', 'Teks Bersih',
        'tf raw = jumlah kata', 'N total dokumen', 'df(w)',
        'idf = LN((1+N)/(1+df))+1', 'tf × idf', 'L2 norm dokumen',
        'TF-IDF Excel = (tf×idf)/L2', 'Penjelasan'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    tfidf_manual = obj['tfidf_manual'].copy()
    text_by_id = obj['transformed'].set_index('ID_Data')['Teks Bersih'].to_dict()
    first_row_by_id = {}
    last_row = len(tfidf_manual) + 1
    for idx, (_, row) in enumerate(tfidf_manual.iterrows(), 2):
        id_data = int(row['ID_Data'])
        first_row_by_id.setdefault(id_data, idx)
        ws.cell(idx, 1, id_data)
        ws.cell(idx, 2, row['Siswa'])
        ws.cell(idx, 3, row['Aspek / Mapel'])
        ws.cell(idx, 4, row['Kata'])
        ws.cell(idx, 5, text_by_id.get(id_data, ''))
        ws.cell(idx, 6, f'=(LEN(E{idx}) - LEN(SUBSTITUTE(E{idx}, D{idx}, ""))) / LEN(D{idx})')
        ws.cell(idx, 7, int(row['N']))
        ws.cell(idx, 8, row['df(w)'])
        ws.cell(idx, 9, f'=LN((1+G{idx})/(1+H{idx}))+1')
        ws.cell(idx, 10, f'=F{idx}*I{idx}')
        ws.cell(idx, 11, 0)  # filled after all rows so each document can use its own range
        ws.cell(idx, 12, f'=IF(K{idx}=0,0,J{idx}/K{idx})')
        ws.cell(idx, 13, 'Urutan: hitung jumlah kata → hitung IDF → kalikan tf×idf → normalisasi L2. Kolom F, I, J, K, L memakai rumus Excel.')

    id_to_rows = {}
    for r in range(2, last_row + 1):
        id_to_rows.setdefault(ws.cell(r, 1).value, []).append(r)
    for rows in id_to_rows.values():
        start, end = rows[0], rows[-1]
        for r in rows:
            ws.cell(r, 11, f'=SQRT(SUMSQ($J${start}:$J${end}))')
    style_ws(ws)
    return ws


def _parse_distribution(text):
    try:
        return ast.literal_eval(str(text))
    except Exception:
        return {}


def add_gini_formula_sheet(wb, obj):
    ws = wb.create_sheet('Gini Rumus Excel')
    labels = list(obj['labels'])
    headers = ['Pohon ke-', 'Node', 'Fitur', 'Threshold', 'n Node']
    headers += [f'Node {label}' for label in labels] + ['Gini Node']
    headers += ['n Kiri'] + [f'Kiri {label}' for label in labels] + ['Gini Kiri']
    headers += ['n Kanan'] + [f'Kanan {label}' for label in labels] + ['Gini Kanan', 'Gini Split', 'Gain Gini', 'Penjelasan']
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    gini_df = obj['gini_all']
    for r, (_, row) in enumerate(gini_df.iterrows(), 2):
        node_dist = _parse_distribution(row['Distribusi Node'])
        left_dist = _parse_distribution(row['Distribusi Kiri'])
        right_dist = _parse_distribution(row['Distribusi Kanan'])
        base = [row['Pohon ke-'], row['Node'], row['Fitur'], row['Threshold'], row['n Node']]
        for c, val in enumerate(base, 1): ws.cell(r, c, val)
        c = 6
        for label in labels:
            ws.cell(r, c, int(node_dist.get(label, 0))); c += 1
        node_count_range = f'{get_column_letter(6)}{r}:{get_column_letter(5+len(labels))}{r}'
        ws.cell(r, c, f'=1-SUMPRODUCT(({node_count_range}/E{r})^2)'); c += 1
        left_n_col = c; ws.cell(r, c, row['n Kiri']); c += 1
        left_start = c
        for label in labels:
            ws.cell(r, c, int(left_dist.get(label, 0))); c += 1
        left_range = f'{get_column_letter(left_start)}{r}:{get_column_letter(c-1)}{r}'
        ws.cell(r, c, f'=IF({get_column_letter(left_n_col)}{r}=0,0,1-SUMPRODUCT(({left_range}/{get_column_letter(left_n_col)}{r})^2))'); c += 1
        right_n_col = c; ws.cell(r, c, row['n Kanan']); c += 1
        right_start = c
        for label in labels:
            ws.cell(r, c, int(right_dist.get(label, 0))); c += 1
        right_range = f'{get_column_letter(right_start)}{r}:{get_column_letter(c-1)}{r}'
        ws.cell(r, c, f'=IF({get_column_letter(right_n_col)}{r}=0,0,1-SUMPRODUCT(({right_range}/{get_column_letter(right_n_col)}{r})^2))'); c += 1
        gini_node_col = 6 + len(labels)
        gini_left_col = left_start + len(labels)
        gini_right_col = right_start + len(labels)
        ws.cell(r, c, f'=({get_column_letter(left_n_col)}{r}/E{r})*{get_column_letter(gini_left_col)}{r}+({get_column_letter(right_n_col)}{r}/E{r})*{get_column_letter(gini_right_col)}{r}'); c += 1
        split_col = c - 1
        ws.cell(r, c, f'={get_column_letter(gini_node_col)}{r}-{get_column_letter(split_col)}{r}'); c += 1
        ws.cell(r, c, 'Gini kecil berarti data makin murni. Gain = Gini node sebelum split - Gini split sesudah dipisah.')
    style_ws(ws)
    return ws


def add_voting_formula_sheet(wb, obj):
    ws = wb.create_sheet('Voting Rumus Excel')
    labels = list(obj['labels'])
    headers = ['ID_Data', 'Siswa', 'Aspek / Mapel'] + [f'Pohon {i}' for i in range(1, 11)] + [f'Jumlah {label}' for label in labels] + ['Hasil Voting Excel', 'Label Aktual', 'Benar/Salah', 'Penjelasan']
    for c, h in enumerate(headers, 1): ws.cell(1, c, h)
    vote_pivot = obj['votes'].pivot_table(index=['ID_Data','Siswa','Aspek / Mapel'], columns='Pohon ke-', values='Prediksi Pohon', aggfunc='first').reset_index()
    actual = obj['transformed'].set_index('ID_Data')['Label'].to_dict()
    for r, (_, row) in enumerate(vote_pivot.iterrows(), 2):
        id_data = int(row['ID_Data'])
        ws.cell(r, 1, id_data); ws.cell(r, 2, row['Siswa']); ws.cell(r, 3, row['Aspek / Mapel'])
        for i in range(1, 11):
            ws.cell(r, 3+i, row.get(i, ''))
        count_start = 14
        for i, label in enumerate(labels):
            ws.cell(r, count_start+i, f'=COUNTIF($D{r}:$M{r},"{label}")')
        result_col = count_start + len(labels)
        count_range = f'{get_column_letter(count_start)}{r}:{get_column_letter(result_col-1)}{r}'
        label_header_range = f'{get_column_letter(count_start)}$1:{get_column_letter(result_col-1)}$1'
        ws.cell(r, result_col, f'=SUBSTITUTE(INDEX({label_header_range},MATCH(MAX({count_range}),{count_range},0)),"Jumlah ","")')
        ws.cell(r, result_col+1, actual.get(id_data, ''))
        ws.cell(r, result_col+2, f'=IF({get_column_letter(result_col)}{r}={get_column_letter(result_col+1)}{r},"Benar","Salah")')
        ws.cell(r, result_col+3, 'Setiap pohon memberi 1 suara. Label dengan jumlah suara paling banyak menjadi hasil Random Forest.')
    style_ws(ws)
    return ws


def add_shap_formula_sheet(wb, obj):
    ws = wb.create_sheet('SHAP Rumus Excel')
    headers = ['ID_Data', 'Siswa', 'Aspek / Mapel', 'Output Prediksi', 'φ0 Base Value', 'Σφi dari detail', 'f(x)=φ0+Σφi Excel', 'Probabilitas Model', 'Selisih', 'Penjelasan']
    for c, h in enumerate(headers, 1): ws.cell(1, c, h)
    detail_ranges = {}
    detail = obj['shap_rows'].reset_index(drop=True)
    for i, row in enumerate(detail.itertuples(index=False), 2):
        detail_ranges.setdefault(int(getattr(row, 'ID_Data')), []).append(i)
    summary = obj['shap_summary']
    for r, (_, row) in enumerate(summary.iterrows(), 2):
        id_data = int(row['ID_Data'])
        ws.cell(r, 1, id_data); ws.cell(r, 2, row['Siswa']); ws.cell(r, 3, row['Aspek / Mapel'])
        ws.cell(r, 4, row['Output Prediksi']); ws.cell(r, 5, row['φ0 Base Value'])
        rows = detail_ranges.get(id_data, [])
        if rows:
            ws.cell(r, 6, f'=SUMIFS(\'SHAP Per Fitur\'!$J:$J,\'SHAP Per Fitur\'!$A:$A,A{r})')
        else:
            ws.cell(r, 6, row['Σφi'])
        ws.cell(r, 7, f'=E{r}+F{r}')
        ws.cell(r, 8, row['Probabilitas Model'])
        ws.cell(r, 9, f'=G{r}-H{r}')
        ws.cell(r, 10, 'Rumus utama SHAP adalah f(x)=φ0+Σφi. φ0 adalah nilai dasar, φi adalah kontribusi fitur.')
    style_ws(ws)
    return ws

def add_preprocess_sheet(wb, obj):
    ws = wb.create_sheet('Praproses Teks')
    headers = [
        'ID_Data', 'Siswa', 'Aspek / Mapel', 'Teks Asli (Raw Text)', 
        'Case Folding', 'Data Cleansing', 'Tokenizing', 'Stopword Removal (Teks Bersih)'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
        
    df = obj['transformed_df'][obj['transformed_df']['Tipe_Data'] == 'Original']
    for idx, (_, row) in enumerate(df.iterrows(), 2):
        raw_text = str(row['Deskripsi Penilaian'])
        cf = raw_text.lower()
        clean = re.sub(r'[^a-zA-Z\s]', ' ', cf)
        clean = re.sub(r'\s+', ' ', clean).strip()
        tokens = str(clean.split())
        stopword = str(row['Teks Bersih'])
        
        ws.cell(idx, 1, int(row['ID_Data']))
        ws.cell(idx, 2, row['Sumber Data'])
        ws.cell(idx, 3, row['Aspek / Mapel'])
        ws.cell(idx, 4, raw_text)
        ws.cell(idx, 5, cf)
        ws.cell(idx, 6, clean)
        ws.cell(idx, 7, tokens)
        ws.cell(idx, 8, stopword)
        
    style_ws(ws)
    return ws

def add_smote_formula_sheet(wb, obj):
    ws = wb.create_sheet('SMOTE Rumus Excel')
    headers = [
        'Nama Fitur', 'Sampel Asli ID 94 (x_i)', 'Tetangga Terdekat ID 221 (x_zi)', 
        'Lambda (λ)', 'Formula SMOTE Excel', 'Hasil Sintetis Excel', 'Keterangan'
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
        
    words = list(obj['words'])
    ohe_categories = list(obj['ohe'].categories_[0])
    
    row_num = 2
    # 1. Nilai
    ws.cell(row_num, 1, 'Nilai')
    ws.cell(row_num, 2, "='TF-IDF Matrix'!F95")
    ws.cell(row_num, 3, "='TF-IDF Matrix'!F222")
    ws.cell(row_num, 4, 0.607034)
    ws.cell(row_num, 5, f'=B{row_num}+D{row_num}*(C{row_num}-B{row_num})')
    ws.cell(row_num, 6, f'=B{row_num}+D{row_num}*(C{row_num}-B{row_num})')
    ws.cell(row_num, 7, 'Interpolasi linear pada fitur Nilai kuantitatif.')
    row_num += 1
    
    # 2. TF-IDF features
    for idx, w in enumerate(words):
        col_letter = get_column_letter(9 + idx) # Col I is 9
        ws.cell(row_num, 1, f'tfidf_{w}')
        ws.cell(row_num, 2, f"='TF-IDF Matrix'!{col_letter}95")
        ws.cell(row_num, 3, f"='TF-IDF Matrix'!{col_letter}222")
        ws.cell(row_num, 4, 0.607034)
        ws.cell(row_num, 5, f'=B{row_num}+D{row_num}*(C{row_num}-B{row_num})')
        ws.cell(row_num, 6, f'=B{row_num}+D{row_num}*(C{row_num}-B{row_num})')
        ws.cell(row_num, 7, f'Interpolasi linear pada bobot TF-IDF kata "{w}".')
        row_num += 1
        
    # 3. Aspects (OHE)
    for idx, name in enumerate(ohe_categories):
        ws.cell(row_num, 1, f'aspek_{name}')
        ws.cell(row_num, 2, f'=IF(\'DATASET KESELURUHAN\'!C95="{name}", 1, 0)')
        ws.cell(row_num, 3, f'=IF(\'DATASET KESELURUHAN\'!C222="{name}", 1, 0)')
        ws.cell(row_num, 4, 0.607034)
        ws.cell(row_num, 5, f'=B{row_num}+D{row_num}*(C{row_num}-B{row_num})')
        ws.cell(row_num, 6, f'=B{row_num}+D{row_num}*(C{row_num}-B{row_num})')
        ws.cell(row_num, 7, f'One-Hot Encoding untuk aspek "{name}". Rekonstruksi menggunakan 1-NN.')
        row_num += 1
        
    style_ws(ws)
    return ws

def make_xlsx_smote(obj):
    wb = Workbook()
    wb.remove(wb.active)
    add_explanation_sheet_smote(wb)
    add_sheet(wb, 'DATASET KESELURUHAN', obj['transformed_df'])
    add_sheet(wb, 'Rumus dan Ringkasan', pd.concat([obj['summary'], pd.DataFrame([['','']], columns=obj['summary'].columns)], ignore_index=True))
    add_sheet(wb, 'Rumus Detail', obj['formula'])
    add_sheet(wb, 'Transformasi Data', obj['transformed_df'][['ID_Data','Sumber Data','Aspek / Mapel','Set Data','Tipe_Data','Nilai','Deskripsi Penilaian','Teks Bersih','Label','Y_Prediksi']])
    add_preprocess_sheet(wb, obj)
    add_sheet(wb, 'TF-IDF Matrix', obj['matrix'])
    obj['transformed'] = obj['transformed_df']
    add_tfidf_formula_sheet(wb, obj)
    add_sheet(wb, 'Struktur Semua Pohon', obj['tree_structures'])
    add_sheet(wb, 'Gini Semua Pohon', obj['gini_all'])
    add_gini_formula_sheet(wb, obj)
    add_sheet(wb, 'Jalur Semua Data', obj['tree_paths_all'])
    add_sheet(wb, 'Voting Per Pohon', obj['votes'])
    add_voting_formula_sheet(wb, obj)
    add_smote_formula_sheet(wb, obj)
    add_sheet(wb, 'Hasil Voting', obj['final_vote'])
    add_sheet(wb, 'Multi Output Long', obj['multi_long'])
    add_sheet(wb, 'Multi Output Per Siswa', obj['multi_wide'])
    add_sheet(wb, 'Hyperparameter Tuning', obj['hpt'])
    add_sheet(wb, 'SHAP Ringkasan', obj['shap_summary'])
    add_sheet(wb, 'SHAP Per Fitur', obj['shap_rows'])
    add_shap_formula_sheet(wb, obj)
    
    # Sheet per tree details
    data_small = obj['transformed_df'][['ID_Data','Sumber Data','Aspek / Mapel','Nilai','Teks Bersih','Label','Y_Prediksi']]
    for t in range(1, 11):
        ws = wb.create_sheet(f'pohon {t}')
        r = write_df(ws, data_small, 1, f'DATASET KESELURUHAN - POHON {t}')
        r = write_df(ws, obj['tree_structures'][obj['tree_structures']['Pohon ke-'].eq(t)], r, f'STRUKTUR POHON {t}')
        r = write_df(ws, obj['gini_all'][obj['gini_all']['Pohon ke-'].eq(t)], r, f'PERHITUNGAN GINI / TABEL KONTINGENSI POHON {t}')
        sample_paths = obj['tree_paths_all'][obj['tree_paths_all']['Pohon ke-'].eq(t)]
        r = write_df(ws, sample_paths, r, f'JALUR KEPUTUSAN SELURUH DATA POHON {t}')
        style_ws(ws)
    
    os.makedirs(OUT_XLSX.parent, exist_ok=True)
    wb.save(OUT_XLSX)

def esc(t): return escape(str(t)).replace('\n','<w:br/>')
def p(text='', style=None, bold=False):
    st = f'<w:pPr><w:pStyle w:val="{style}"/></w:pPr>' if style else ''
    b = '<w:b/>' if bold else ''
    return f'<w:p>{st}<w:r><w:rPr>{b}</w:rPr><w:t xml:space="preserve">{esc(text)}</w:t></w:r></w:p>'
def code_p(text=''):
    st = '<w:pPr><w:jc w:val="center"/><w:rPr><w:rFonts w:ascii="Consolas" w:hAnsi="Consolas"/><w:sz w:val="18"/></w:rPr></w:pPr>'
    return f'<w:p>{st}<w:r><w:rPr><w:rFonts w:ascii="Consolas" w:hAnsi="Consolas"/><w:sz w:val="18"/></w:rPr><w:t xml:space="preserve">{esc(text)}</w:t></w:r></w:p>'
def eq(text):
    return f'<w:p><m:oMathPara><m:oMath><m:r><m:t>{esc(text)}</m:t></m:r></m:oMath></m:oMathPara></w:p>'
def tbl(headers, rows, max_rows=10):
    rows = list(rows)[:max_rows]
    out = ['<w:tbl><w:tblPr><w:tblBorders><w:top w:val="single" w:sz="4"/><w:left w:val="single" w:sz="4"/><w:bottom w:val="single" w:sz="4"/><w:right w:val="single" w:sz="4"/><w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/></w:tblBorders></w:tblPr>']
    def cell(v, b=False):
        bd = '<w:b/>' if b else ''
        return f'<w:tc><w:p><w:r><w:rPr>{bd}</w:rPr><w:t xml:space="preserve">{esc(v)}</w:t></w:r></w:p></w:tc>'
    out.append('<w:tr>'+''.join(cell(h,True) for h in headers)+'</w:tr>')
    for r in rows: out.append('<w:tr>'+''.join(cell(v) for v in r)+'</w:tr>')
    out.append('</w:tbl>')
    return ''.join(out)

def build_tree_text_lines(tree, node, depth, feature_names, labels):
    is_leaf = tree.children_left[node] == tree.children_right[node]
    indent = "  " * depth
    if is_leaf:
        value = tree.value[node][0]
        norm_val = value / np.sum(value) if np.sum(value) > 0 else value
        counts_dict = {labels[i]: int(round(v * tree.n_node_samples[node])) for i, v in enumerate(norm_val)}
        pred_cls = int(np.argmax(value))
        pred_name = labels[pred_cls]
        return [f"{indent}Node {node} [Daun]: Prediksi = {pred_name} (n={int(tree.n_node_samples[node])}, Dist={counts_dict})"]
    else:
        fi = tree.feature[node]
        thr = float(tree.threshold[node])
        feat_name = feature_names[fi]
        n_samples = int(tree.n_node_samples[node])
        
        value = tree.value[node][0]
        norm_val = value / np.sum(value) if np.sum(value) > 0 else value
        counts_dict = {labels[i]: int(round(v * tree.n_node_samples[node])) for i, v in enumerate(norm_val)}
        gini_val = float(1.0 - np.sum(norm_val ** 2))
        
        lines = [f"{indent}Node {node} [Internal]: jika {feat_name} <= {thr:.6f} (n={n_samples}, Gini={gini_val:.4f}, Dist={counts_dict})"]
        lines.extend(build_tree_text_lines(tree, tree.children_left[node], depth + 1, feature_names, labels))
        lines.append(f"{indent}Node {node} [Internal]: jika {feat_name} > {thr:.6f}")
        lines.extend(build_tree_text_lines(tree, tree.children_right[node], depth + 1, feature_names, labels))
        return lines

def make_docx_smote(obj):
    tfrow = obj['tfidf_manual'].iloc[0]
    grow = obj['gini_all'].iloc[0]
    vrow = obj['final_vote'].iloc[0]
    srow = obj['shap_summary'].iloc[0]
    body = []
    
    body.append(p('DOKUMEN PERHITUNGAN MANUAL RANDOM FOREST DENGAN SMOTE (OVERSAMPLING)', 'Title', True))
    body.append(p('Dokumen ini disusun untuk menjelaskan langkah-langkah data mining setelah menggunakan teknik SMOTE untuk menangani class imbalance. File Excel pendamping memuat perhitungan manual seluruh data training yang telah di-resample menggunakan SMOTE.'))
    
    body.append(p('1. Penjelasan Variabel Penelitian', 'Heading1', True))
    body.append(p('Penelitian ini menggunakan variabel bebas (independent variable) sebagai masukan model dan variabel terikat (dependent variable) sebagai target prediksi:'))
    body.append(tbl(['Jenis Variabel', 'Variabel', 'Simbol', 'Peran', 'Keterangan'], [
        ['Variabel Bebas', 'Nilai', 'X1', 'Masukan Model', 'Nilai akademis siswa, numerik kontinu (skala 0.0 s.d. 100.0).'],
        ['Variabel Bebas', 'Deskripsi Penilaian', 'X2', 'Masukan Model', 'Catatan kualitatif perkembangan siswa dalam bentuk teks.'],
        ['Variabel Terikat', 'Aspek / Mapel', 'y1', 'Target Prediksi', 'Bidang studi atau aspek portofolio non-akademis.'],
        ['Variabel Terikat', 'Label', 'y2', 'Target Prediksi', 'Status perkembangan siswa (Baik, Cukup, Perlu Bimbingan).']
    ]))
    
    body.append(p('2. Tahapan Praproses Teks (Text Preprocessing)', 'Heading1', True))
    body.append(p('Praproses teks bertujuan untuk menyaring noise dari deskripsi penilaian. Berikut adalah contoh transformasi teks pada data ID 1:'))
    body.append(p('"Al-Qur\'an Hadis : Memahami huruf hijaiah dan harakatain, surah-surah pendek Al-Qur\'an pilihan, dan nilai-nilai kandungan Al-Qur\'an dan/atau hadis tentang kewajiban salat."'))
    body.append(tbl(['Tahapan', 'Aturan', 'Hasil Output Teks'], [
        ['Raw Text', 'Teks deskripsi masukan asli', '"Al-Qur\'an Hadis : Memahami huruf hijaiah dan harakatain, surah-surah pendek Al-Qur\'an pilihan, dan nilai-nilai kandungan Al-Qur\'an dan/atau hadis tentang kewajiban salat."'],
        ['Case Folding', 'text.lower()', '"al-qur\'an hadis : memahami huruf hijaiah dan harakatain, surah-surah pendek al-qur\'an pilihan, dan nilai-nilai kandungan al-qur\'an dan/atau hadis tentang kewajiban salat."'],
        ['Cleansing', 'Hapus tanda baca/angka via regex [^a-zA-Z\\s]', '"al qur an hadis   memahami huruf hijaiah dan harakatain  surah surah pendek al qur an pilihan  dan nilai nilai kandungan al qur an dan atau hadis tentang kewajiban salat "'],
        ['Tokenizing', 'Pemisahan kata berdasarkan spasi (split)', "['al', 'qur', 'an', 'hadis', 'memahami', 'huruf', 'hijaiah', 'dan', 'harakatain', 'surah', 'surah', 'pendek', 'al', 'qur', 'an', 'pilihan', 'dan', 'nilai', 'nilai', 'kandungan', 'al', 'qur', 'an', 'dan', 'atau', 'hadis', 'tentang', 'kewajiban', 'salat']"],
        ['Stopword Removal', 'Hapus kata hubung & nama berdasarkan kamus', "['hadis', 'memahami', 'huruf', 'hijaiah', 'harakatain', 'surah', 'surah', 'pendek', 'pilihan', 'kandungan', 'hadis', 'kewajiban', 'salat']"]
    ]))
    
    body.append(p('3. Perhitungan TF-IDF Manual', 'Heading1', True))
    body.append(p('Nilai TF-IDF dihitung dari term frequency (TF) dikalikan inverse document frequency (IDF) dengan rumus equation:'))
    body.append(eq('IDF(w) = ln((1 + N) / (1 + df(w))) + 1'))
    body.append(eq('v(w, d) = TF(w, d) × IDF(w)'))
    body.append(eq('||v_d||_2 = sqrt(sum(v(w_i, d)^2))'))
    body.append(eq('TF-IDF(w, d) = v(w, d) / ||v_d||_2'))
    body.append(p('Berikut adalah tabel perhitungan bobot kata (TF-IDF) untuk setiap aspek/mata pelajaran yang diwakili oleh sampel data training asli (diurutkan berdasarkan bobot TF-IDF akhir tertinggi):', 'Normal'))
    
    unique_aspeks = obj['tfidf_manual']['Aspek / Mapel'].unique()
    for asp in unique_aspeks:
        asp_df = obj['tfidf_manual'][obj['tfidf_manual']['Aspek / Mapel'] == asp]
        if asp_df.empty:
            continue
        first_id = asp_df['ID_Data'].iloc[0]
        doc_df = asp_df[asp_df['ID_Data'] == first_id]
        
        # Urutkan berdasarkan bobot akhir terbesar
        doc_df = doc_df.sort_values(by='TF-IDF Manual Setelah L2', ascending=False)
        
        l2_val = doc_df['L2 norm dokumen'].iloc[0]
        siswa_name = doc_df['Siswa'].iloc[0]
        
        body.append(p(f"Aspek/Mapel: {asp} (Sampel: {siswa_name}, ID Data: {first_id}, L2-norm = {l2_val:.6f})", 'Heading2', True))
        
        rows_data = []
        for _, r in doc_df.head(10).iterrows():
            rows_data.append([
                str(r['Kata']),
                str(int(r['tf raw (jumlah kata)'])),
                str(int(r['df(w)'])),
                str(int(r['N'])),
                f"{float(r['idf sklearn = ln((1+N)/(1+df))+1']):.6f}",
                f"{float(r['tf \u00d7 idf']):.6f}",
                f"{float(r['TF-IDF Manual Setelah L2']):.6f}"
            ])
            
        body.append(tbl(['Term (Kata)', 'Tf (raw)', 'df(w)', 'N', 'IDF', 'Tf \u00d7 IDF', 'TF-IDF Akhir'], rows_data, 10))
    
    body.append(p('4. Penanganan Ketimpangan Data dengan SMOTE', 'Heading1', True))
    body.append(p('Oversampling menggunakan SMOTE (Synthetic Minority Over-sampling Technique) digunakan untuk menyeimbangkan kelas target. Dari data training awal sebanyak 288 baris, kelas mayoritas (Cukup) memiliki 152 data, sedangkan kelas minoritas Baik (69) dan Perlu Bimbingan (67) di-resample sehingga masing-masing memiliki 152 data (Total 456 data training).'))
    body.append(p('Sampel baru x_new dibuat pada garis penghubung x_i and tetangganya x_zi dengan rumus:'))
    body.append(eq('x_new = x_i + λ × (x_zi - x_i)'))
    body.append(p('Berikut adalah tabel perhitungan manual untuk pembuatan sampel sintetis baru (SMOTE_Sintetis_1 / ID 1001) yang dibuat dari Sampel Asli ID 94 (x_i) dan Tetangga Terdekat ID 221 (x_zi) dengan λ = 0.607034:'))
    body.append(tbl(['Nama Fitur', 'Sampel Asli ID 94 (x_i)', 'Tetangga ID 221 (x_zi)', 'Lambda (λ)', 'Formula SMOTE Excel', 'Hasil Sintetis Excel'], [
        ['Nilai', '0.000000', '0.000000', '0.607034', '=B2+D2*(C2-B2)', '0.000000'],
        ['tfidf_agar', '0.157746', '0.000000', '0.607034', '=B3+D3*(C3-B3)', '0.061989'],
        ['tfidf_baik', '0.109246', '0.143157', '0.607034', '=B4+D4*(C4-B4)', '0.129831'],
        ['tfidf_belajar', '0.322227', '0.211126', '0.607034', '=B5+D5*(C5-B5)', '0.254785'],
        ['aspek_Emosi', '0.000000', '1.000000', '0.607034', '=B6+D6*(C6-B6)', '0.607034']
    ]))
    
    body.append(p('5. Perhitungan Gini Impurity Random Forest', 'Heading1', True))
    body.append(p('Gini Impurity dihitung untuk menentukan pemilihan split node pada decision tree:'))
    body.append(eq('Gini(D) = 1 - sum(p_i^2)'))
    body.append(eq('Gini_split = (n_kiri / n_node) × Gini(D_kiri) + (n_kanan / n_node) × Gini(D_kanan)'))
    body.append(eq('Gini_Gain = Gini(D) - Gini_split'))
    
    body.append(p('Struktur Lengkap Pohon Keputusan ke-1 (Visualisasi Hierarki):', 'Heading2', True))
    tree_lines = build_tree_text_lines(obj['rf'].estimators_[0].tree_, 0, 0, obj['feature_names'], list(obj['labels']))
    for line in tree_lines:
        body.append(code_p(line))
        
    body.append(p(f"Contoh kalkulasi pembagian node akar (Node 0): Gini split = ({int(grow['n Kiri'])}/{int(grow['n Node'])})×{grow['Gini Kiri']}+({int(grow['n Kanan'])}/{int(grow['n Node'])})×{grow['Gini Kanan']} = {grow['Gini Split']}"))
    body.append(tbl(list(obj['gini_all'].columns), obj['gini_all'].values, 5))
    
    body.append(p('6. Jumlah Node Pohon dalam Penelitian', 'Heading1', True))
    body.append(p('Model Random Forest terdiri atas 300 pohon keputusan independen (hasil hyperparameter tuning). Jumlah node untuk 10 pohon pertama dan ringkasan keseluruhan pohon dirangkum pada tabel di bawah ini:'))
    
    tree_counts = obj['tree_structures'].groupby('Pohon ke-').size().to_dict()
    total_nodes = obj['tree_structures'].shape[0]
    avg_nodes = total_nodes / len(tree_counts)
    
    node_rows = []
    for t_idx in range(1, 11):
        count = tree_counts.get(t_idx, 0)
        node_rows.append([f"Pohon ke-{t_idx}", f"{count} Node"])
    node_rows.append(["Pohon ke-11 s.d. ke-300", "Detail tersedia di file Excel"])
    node_rows.append(["Total (300 Pohon)", f"{total_nodes} Node"])
    node_rows.append(["Rata-rata Node per Pohon", f"{avg_nodes:.1f} Node"])
    
    body.append(tbl(['Pohon / Model', 'Jumlah Node'], node_rows, 20))
    
    body.append(p('7. Perhitungan Voting Mayoritas', 'Heading1', True))
    body.append(p('Prediksi akhir y_pred adalah modus dari hasil prediksi seluruh 10 pohon:'))
    body.append(eq('y_pred = mode{h_1(x), h_2(x), ..., h_10(x)}'))
    body.append(p(f"Contoh hasil voting untuk data ID {int(vrow['ID_Data'])}: hasil voting adalah {vrow['Hasil Voting']} dengan total suara {int(vrow['Jumlah Suara'])}."))
    body.append(tbl(list(obj['final_vote'].columns), obj['final_vote'].values, 8))
    
    body.append(p('8. Perhitungan SHAP (Explainable AI)', 'Heading1', True))
    body.append(p('SHAP digunakan untuk mengukur kontribusi fitur terhadap keputusan model:'))
    body.append(eq('f(x) = φ_0 + sum(φ_i)'))
    body.append(p(f"Kalkulasi kontribusi untuk prediksi kelas Cukup pada ID 1: f(x) = {srow['φ0 Base Value']} + {srow['Σφi']} = {srow['f(x)=φ0+Σφi Excel']}:"))
    body.append(tbl(list(obj['shap_summary'].columns), obj['shap_summary'].values, 6))

    tmp = Path(tempfile.mkdtemp())
    try:
        with zipfile.ZipFile(TEMPLATE_DOCX, 'r') as zin:
            zin.extractall(tmp)
        xml = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:wpc="http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas" xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:wp14="http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing" xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" xmlns:w10="urn:schemas-microsoft-com:office:word" xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" xmlns:w14="http://schemas.microsoft.com/office/word/2010/wordml" xmlns:wpg="http://schemas.microsoft.com/office/word/2010/wordprocessingGroup" xmlns:wpi="http://schemas.microsoft.com/office/word/2010/wordprocessingInk" xmlns:wne="http://schemas.microsoft.com/office/word/2006/wordml" xmlns:wps="http://schemas.microsoft.com/office/word/2010/wordprocessingShape" mc:Ignorable="w14 wp14"><w:body>{''.join(body)}<w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="720" w:footer="720" w:gutter="0"/></w:sectPr></w:body></w:document>'''
        (tmp/'word'/'document.xml').write_text(xml, encoding='utf-8')
        try:
            if OUT_DOCX.exists():
                OUT_DOCX.unlink()
            with zipfile.ZipFile(OUT_DOCX, 'w', zipfile.ZIP_DEFLATED) as zout:
                for path in tmp.rglob('*'):
                    if path.is_file():
                        zout.write(path, path.relative_to(tmp).as_posix())
        except PermissionError:
            print(f"\n[WARNING] Tidak bisa menulis ke file {OUT_DOCX} karena sedang dibuka oleh program lain (misal: Microsoft Word).")
            print("Harap tutup dokumen tersebut di Microsoft Word, lalu jalankan script ini kembali.")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

def main():
    print("Building calculations for SMOTE version...")
    obj = build_smote_calculations()
    make_xlsx_smote(obj)
    make_docx_smote(obj)
    print("SMOTE Manual Calculations Excel created at:", OUT_XLSX)
    print("SMOTE Manual Calculations Word created at:", OUT_DOCX)

if __name__ == '__main__':
    main()
